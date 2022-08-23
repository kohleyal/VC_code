import os
import csv
import tkinter as tk 
from tkinter import ttk
from tkinter import filedialog
from datetime import datetime
from openpyxl import load_workbook
from lib.SA_tank_class import Sartorious_tank



class MainApp(ttk.Notebook):
    def __init__(self, parent, *args, **kwargs):
        ttk.Notebook.__init__(self, parent, *args, **kwargs)
        self.parent = parent
        self.notebook = ttk.Notebook(self.parent)
        self.notebook.grid(row=0, column=0, sticky=tk.N+tk.E+tk.S+tk.W)
        self.vessel_select_tab_frame = Vessel_Select_tab(self.notebook)
        self.offgas_tab_frame = Offgas_tab(self.notebook)
        self.combine_online_run_offgas_tab = CombineAll_tab(self.notebook)
        self.daily_sample_tab = Daily_Sample_Tab(self.notebook)

        self.vessel_select_tab_frame.columnconfigure(0, weight=1) #make the tab fill and stretch with the window

        self.notebook.add(self.vessel_select_tab_frame, text='Vessel Select')
        self.notebook.add(self.offgas_tab_frame, text='OffGas')
        self.notebook.add(self.combine_online_run_offgas_tab, text='Combine All')
        self.notebook.add(self.daily_sample_tab, text='Daliy Sample')


class Vessel_Select(ttk.Frame):
    '''Base template which each GUI notebook tab is built off of.'''
    def __init__(self, parent, *args, **kwargs):
        ttk.Frame.__init__(self, parent, *args, **kwargs)
        #-----Vessel Selection Widgets --------#
        self.vessel_select_frame = ttk.LabelFrame(self, padding=5, relief='groove')
        self.vessel_info_frame = ttk.LabelFrame(self, text='Instructions')
        self.vessel_select_frame.grid(row=0 , column=0, padx=10, pady=10, sticky=tk.NW)
        self.vessel_select_frame.columnconfigure(1,weight=1, minsize=400)

        self.vessel_info_canvas = tk.Canvas(self.vessel_info_frame, width=200)
        self.vessel_info_canvas.grid(row=0, column=0)
        self.vessel_info_frame.grid(row=0, column=1, rowspan=3, padx=10, pady=10, sticky= tk.N+tk.S+tk.W)

        self.make_buttons()

        self.v1 = Sartorious_tank('V1')
        self.v2 = Sartorious_tank('V2')
        self.v3 = Sartorious_tank('V3')
        self.v4 = Sartorious_tank('V4')
        self.v5 = Sartorious_tank('V5')

    def make_buttons(self):
        #set file path str varaible for each vessel
        self.v1_file_text = tk.StringVar()
        self.v1_file_text.set('V1 File')
        self.v2_file_text = tk.StringVar()
        self.v2_file_text.set('V2 File')
        self.v3_file_text = tk.StringVar()
        self.v3_file_text.set('V3 File')
        self.v4_file_text = tk.StringVar()
        self.v4_file_text.set('V4 File')
        self.v5_file_text = tk.StringVar()
        self.v5_file_text.set('V5 File')

        #create the button and labels for the vessels 
        self.v1_bttn = ttk.Button(self.vessel_select_frame, text='Vessel 1', command=lambda: self.get_file_path(self.v1_file_text))
        self.v2_bttn = ttk.Button(self.vessel_select_frame, text='Vessel 2', command=lambda: self.get_file_path(self.v2_file_text))
        self.v3_bttn = ttk.Button(self.vessel_select_frame, text='Vessel 3', command=lambda: self.get_file_path(self.v3_file_text))
        self.v4_bttn = ttk.Button(self.vessel_select_frame, text='Vessel 4', command=lambda: self.get_file_path(self.v4_file_text))
        self.v5_bttn = ttk.Button(self.vessel_select_frame, text='Vessel 5', command=lambda: self.get_file_path(self.v5_file_text))
        self.v1_file = ttk.Label(self.vessel_select_frame, wraplength=420, textvariable=self.v1_file_text)
        self.v2_file = ttk.Label(self.vessel_select_frame, wraplength=420, textvariable=self.v2_file_text)
        self.v3_file = ttk.Label(self.vessel_select_frame, wraplength=420, textvariable=self.v3_file_text)
        self.v4_file = ttk.Label(self.vessel_select_frame, wraplength=420, textvariable=self.v4_file_text)
        self.v5_file = ttk.Label(self.vessel_select_frame, wraplength=420, textvariable=self.v5_file_text)

        #place buttons
        self.v1_bttn.grid(row=0, column=0, padx=4, pady=4, sticky= tk.W)
        self.v2_bttn.grid(row=1, column=0, padx=4, pady=4, sticky= tk.W)
        self.v3_bttn.grid(row=2, column=0, padx=4, pady=4, sticky= tk.W)
        self.v4_bttn.grid(row=3, column=0, padx=4, pady=4, sticky= tk.W)
        self.v5_bttn.grid(row=4, column=0, padx=4, pady=4, sticky= tk.W)
        self.v1_file.grid(row=0, column=1, padx=10)
        self.v2_file.grid(row=1, column=1, padx=10)
        self.v3_file.grid(row=2, column=1, padx=10)
        self.v4_file.grid(row=3, column=1, padx=10)
        self.v5_file.grid(row=4, column=1, padx=10)

    def get_file_path(self, file_text):
        file_path = filedialog.askopenfilename(filetypes= [('CSV files','.csv'), ('Any','.*')])
        #tkinter set text on GUI
        file_text.set(file_path)
        return 

  
class Vessel_Select_tab(Vessel_Select):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        
        self.batchID_frame = ttk.LabelFrame(self, padding=5, relief='groove', text='Enter Batch ID')
        self.batchID_entry = ttk.Entry(self.batchID_frame) 
        self.combine_tanks_bttn = ttk.Button(self, text= 'Combine Tanks', command=self.combine_tank_files)

        self.batchID_frame.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)
        self.batchID_entry.grid(padx=4)
        self.combine_tanks_bttn.grid(row= 2, column=0, padx=20, pady=5, sticky= tk.W)

        self.vessel_select_instructions = '''Enter raw run sheets exported from the Sartorius software. Even single vessel runs\n\nEnter the batch number.\n\nClick Combine Tanks'''        
        self.vessel_info_text_id = self.vessel_info_canvas.create_text(20, 40, text=self.vessel_select_instructions, anchor=tk.NW, justify='center', width=180, font=('Times', '11'))

        self.vessel_select_frame.config( text = 'Vessel MFCS Run Data')


    def combine_tank_files(self):
        # Setup for transforing files.   #possibly put this in a text file and make editing it part of the gui.
        v1order = [[1, 2, 3, 4, 9, 10, 11, 22, 41, 8, 15, 20, 23, 32, 33,37, 45],'V1']             # old SA sftware output [1, 2, 3, 4, 9, 10, 11, 22, 41, 8, 14, 18, 23, 29, 36, 39, 44]
        v2order = [[1, 2, 3, 4, 9, 10, 11, 22, 41, 7, 15, 19, 23, 30, 34, 39, 45],'V2']            # old SA sftware output [1, 2, 3, 4, 9, 10, 11, 22, 41, 7, 14, 21, 24, 29, 34, 40, 42]
        v3order = [[1, 2, 3, 4, 5, 6, 7, 14, 33, 34, 8, 12, 16, 19, 21, 25, 27, 31, 35, 40],'V3']  # old SA sftware output [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19] 
        v4order = [[1, 2, 3, 4, 5, 6, 7, 14, 33, 34, 9, 12, 15, 19, 23, 25, 29, 32, 35, 38],'V4']  # old SA sftware output [0,1,2,3,4,18,19,5,6,7,8,9,10,11,12,13,14,15,16,17]
        v5order = [[1, 2, 3, 4, 5, 6, 7, 20, 30, 31, 11, 13, 19, 21, 25, 29, 33, 35, 38, 41],'V5'] # old SA sftware output [0,1,2,3,4,5,6,10,11,12,13,14,15,16,17,18,19,20,21,22]
        v1_file_path = self.v1_file_text.get()
        v2_file_path = self.v2_file_text.get()
        v3_file_path = self.v3_file_text.get()
        v4_file_path = self.v4_file_text.get()
        v5_file_path = self.v5_file_text.get()
        batchID = self.batchID_entry.get()
        #check if there is a batch number entered.
        if not batchID:
            print('No Batch ID entered') #TODO add to gui in future.
            return 
        combin_dic = {v1_file_path:v1order, v2_file_path:v2order, v3_file_path:v3order, v4_file_path:v4order, v5_file_path:v5order}

        file_header = ['Batch ID', 'PDatTime', 'Age (h)', 'ACID (mL)', 'AFOM (mL)', 'BASE (mL)', 'EXT_1 (%)', 'EXT_2 (%)', 'O2_T (L)', 'SUTA (mL)',
        'SUTB (mL)', 'GASF (l/m)', 'JTMP (°C)', 'O2EN (%)', 'pH', 'pO2 (%)', 'STIR (rpm)', 'SUBA(%)', 'SUBB (%)', 'TEMP (°C)', 'VWGH (kg)'] # the first row headers for file

        #--------------------File transformation code-----------------------#

        #takes the batch number prefix and dictionary of file name and order. Opens each, reorders, and saves under one file. 
        newfile = [] # open each file and add to this list for each vessel used. 
        newfile.append(file_header)
        for sa_file in combin_dic:
            vessel = combin_dic.get(sa_file)[1]
            if sa_file.lower().endswith('.csv'):
                with open(sa_file,"r") as SAfile:
                    #V2 file has a different output than the other two. 
                    # if vessel == 'V2' or vessel == 'V1':
                    reader = csv.reader(SAfile, delimiter=';')
                    # else:
                    #     reader = csv.reader(SAfile) #old SA software output needed this for the older towers. Keep for now if needed to use on again. 
                    for row in reader:
                        if row:
                            row = [row[i] for i in combin_dic.get(sa_file)[0]]  # takes the row in ls form and changes the ordered to the past in list order.
                            if vessel == 'V2' or vessel == 'V1':    #put in blank cells for V2 and V1. Cant use above to move values that are not there.  
                                row.insert(9,'')
                                row.insert(17,'')
                                row.append('') #make V1 and V2 as long as V3-V5
                            row.insert(0,'') # put in column for batchID
                            if reader.line_num > 3: #skip header in first three rows of data file. (>3, can reverse to get only headers)
                                row[0] = batchID+ ' ' +vessel
                                newfile.append(row)  
                print(f'{vessel} file was transformed') #TODO add to gui
            else:
                print('No CSV file entered for ' + vessel) #TODO add to gui
        file_name = filedialog.asksaveasfilename(filetypes= [('CSV files','.csv')], defaultextension='.csv') #put in new file name.
        with open(file_name, 'w', newline='') as file: 
            writer = csv.writer(file)
            for row in newfile:
                writer.writerow(row)
        print("New file saved!") #TODO add to gui later


class Offgas_tab(Vessel_Select):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)

        self.offgas_combine_bttn = ttk.Button(self, text='Combine Offgas', command=self.offgas_combine)
        self.offgas_combine_bttn.grid(row=1, column=0, padx=20, pady=5, sticky=tk.W)

        self.vessel_select_instructions = '''Combines all offgas csv files into one file.\n\nSelect files with offgas datapoints within desired timerange.\n\nClick Offgas button'''
        self.vessel_info_text_id = self.vessel_info_canvas.create_text(20, 40, text=self.vessel_select_instructions, anchor=tk.NW, justify='center', width=180, font=('Times', '11'))

        self.vessel_select_frame.config( text = 'OffGas Folders')
        self.v1_bttn.config(command=lambda: self.get_folder_path(self.v1_file_text, self.v1))
        self.v2_bttn.config(command=lambda: self.get_folder_path(self.v2_file_text, self.v2))
        self.v3_bttn.config(command=lambda: self.get_folder_path(self.v3_file_text, self.v3))
        self.v4_bttn.config(command=lambda: self.get_folder_path(self.v4_file_text, self.v4))
        self.v5_bttn.config(command=lambda: self.get_folder_path(self.v5_file_text, self.v5))
        self.v1_file_text.set('V1 Offgas Folder')
        self.v2_file_text.set('V2 Offgas Folder')
        self.v3_file_text.set('V3 Offgas Folder')
        self.v4_file_text.set('V4 Offgas Folder')
        self.v5_file_text.set('V5 Offgas Folder')

    def offgas_combine(self):
        #header set already in list below. Modified some titles to not to use special characters.
        self.offgas_combined = [["Tank", "Date_Time", "N2", "O2", "Ar", "CO2", "RMS Flow", "p.CDC", "p.OXC", "RQ", "mass28", "mass32", "mass40", "mass44"]]

        self.tank_list = [tank for tank in [self.v1, self.v2, self.v3, self.v4, self.v5] if len(tank.offgas_folder) > 0 ]

        #get list of csv files to combine
        for tank in self.tank_list: 
            tank_list = []
            for off_gas_file in tank.offgas_folder:
                with open(off_gas_file, 'r') as off_gas:
                    reader = csv.reader(off_gas)
                    for line_index, line in enumerate(reader):
                        if line_index > 0:
                            line.insert(0, tank.tank_id)
                            #minus one because there is a trailing space on the offgas output.
                            tank_list.append(line[:len(line)-1])
            self.offgas_combined.extend(sorted(tank_list, reverse = False, key = self.sort_by_time))
        
        #save the compiled data to one csv file. may make this a return statement.
        print(f'Writing to file.')
        save_file = filedialog.asksaveasfilename(filetypes= [('CSV files','.csv')], defaultextension='.csv') 
        with open(f'{save_file}', 'w', newline='') as file:
            writer = csv.writer(file)
            for offgas_line in self.offgas_combined:
                writer.writerow(offgas_line)
        print('Saved')
    
    def sort_by_time(self, list_item):
        return datetime.strptime(list_item[1],"%Y/%m/%d %H:%M:%S.%f")

    def get_folder_path(self, file_text, tank):
        '''set tank object offgas folder value to a tuple of file names to compile together'''
        tank.offgas_folder = filedialog.askopenfilenames(filetypes= [('CSV files','.csv'), ('Any','.*')])
        #tkinter set text on GUI
        if len(tank.offgas_folder) > 6:
            file_text.set(f'{tank.offgas_folder[0]}....{tank.offgas_folder[-1]}')
        else: 
            file_text.set(tank.offgas_folder)
        return 
    
    def walk_dirctories(self, directory):
        #removed use but keep for possible future use
        #get folder to crawl
        #returns list of reactor system and file path
        file_list  = []
        def crabwalk (currD) :
            #get contents of directory
            scan = os.scandir(currD)
            for fle in scan:
                if fle.is_dir() :
                    #if there is another directory. Recursive down into it. 
                    crabwalk(fle.path)
                else:
                    file_list.append(fle.path)

        crabwalk(directory)
        return file_list


class HC :
    '''Convert into other tank class or into the online offgas offline class
        leftover from first file combine script and hold hardcode inputs for file combination.'''
    #offline HC
    #offline test conditions data locations
    offline_tc_col = [1,2,3,5]
    offline_tc_row = [5,6,7,8,9]

    #offline data location parameters
    offline_eft_col = 3
    offline_row_start = 12
    offline_column_start = 11
    
    #offline headers location parameters
    offline_header_row_start =11
    offline_header_row_end = 11
    offline_header_column_start = 11

    #online HC
    #vessel id column index in row
    online_id_column = 0
    
    #test conditinos placement in row
    online_tc_col_start = 3

    #onilne eft index in row
    online_eft_index = 2


class CombineAll_tab(Vessel_Select):
    '''Combine compiled raw tank file, compiled offgas file, and run sheet'''
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.vessel_select_frame.config( text = 'Final Combination Files')
        self.on_off_file_bttn = ttk.Button(self, text='Combine Online, Offgas, Offline', command=self.combine_online_offgas_runsheet_buttonCmd)
        self.on_off_file_bttn.grid(row=1, column=0, padx=10, pady=5, sticky=tk.W)

        
        self.vessel_select_instructions = '''Enter each file.\n\nShould work with online file and just a offgas or offline file.\n\nClick Combine button'''
        self.vessel_info_text_id = self.vessel_info_canvas.create_text(20, 40, text=self.vessel_select_instructions, anchor=tk.NW, justify='center', width=180, font=('Times', '11'))

        self.v1_bttn.config(text='Online File')
        self.v1_file_text.set('Online File')
        self.v2_bttn.config(text='Offgas File')
        self.v2_file_text.set('Offgas File')
        self.v3_bttn.config(text='Offline File', command=lambda: self.get_file_path_xlsm(self.v3_file_text))
        self.v3_file_text.set('Offline File')
        self.v4_bttn.destroy()
        self.v5_bttn.destroy()
        self.v4_file.destroy()
        self.v5_file.destroy()


    def combine_online_offgas_runsheet_buttonCmd(self):
        min_time, max_time, tank_file = self.min_max_timestamp( time_column=1) #TODO move hardcode to top or class.
        if self.v2_file_text.get().endswith('.csv'):
            self.online_offgas_tank_file = self.input_offgas(tank_file, min_time, max_time, 1) #TODO move hardcode to top or class.
        else:
            self.online_offgas_tank_file = tank_file
        if self.v3_file_text.get().endswith('.xlsm'):
            self.online_offgas_tank_file = self.combine_onlineOffgas_and_offline()

        #ask for a file name and write a new csv with the filename. 
        file_name = filedialog.asksaveasfilename(filetypes= [('CSV files','.csv')], defaultextension='.csv') #get name for compiled file.
        with open(file_name, 'w', newline='') as file: 
            writer = csv.writer(file)
            for row in self.online_offgas_tank_file:
                writer.writerow(row)
        print("New file saved!")         


    def min_max_timestamp(self, time_column):
        '''Get the min and max time from the online file'''
        total_file =[]
        with open(self.v1_file_text.get(), 'r') as file:
            read_file = csv.reader(file)
            min_time = datetime(3000, 1, 1)
            max_time = datetime(1, 1, 1)
            for index, line in enumerate(read_file):
                if index > 0:
                    time = datetime.strptime(line[time_column],"%m/%d/%Y %I:%M:%S %p")
                    line[time_column] = time
                    if time < min_time:
                        min_time = time
                    if time > max_time:
                        max_time = time
                total_file.append(line)
            return [min_time, max_time, total_file]


    def input_offgas(self, tank_file, min_time, max_time, time_column):
        '''Add compiled offgas data to compiled online data'''
        #TODO check if hard code numbers should stay or go.
        with open(self.v2_file_text.get(), 'r') as offgas_file:
            offgas = csv.reader(offgas_file)
            tank_file_index = 1
            for index, offgas_line in enumerate(offgas):
                if index < 1:
                    header = offgas_line #need to put header in but only once
                    insertion_length_list = ['' for x in header[2:]] 
                    total_line_length = len(tank_file[index]) + len(header[2:])
                    continue
                offgas_time = datetime.strptime(offgas_line[time_column],"%Y/%m/%d %H:%M:%S.%f")
                if offgas_time < min_time:
                    continue
                elif offgas_time > max_time:
                    #if the offgas time is past the run time skip
                    continue
                for tank_line in range(tank_file_index, len(tank_file)):
                    #check that the tank id's match
                    if  offgas_line[0].lower() not in tank_file[tank_line][0].lower():
                        continue
                    if tank_file[tank_line][1] > offgas_time:
                        #if we've already inserted a line here. Do not exstend it farther. Skip.
                        if len(tank_file[tank_line]) >= total_line_length:
                            break
                        #if tank time is greater than offgas. put in the offgas information.
                        #then start loop again at next line. 
                        tank_file[tank_line].extend(offgas_line[2:])
                        tank_file_index = tank_line 
                        break
                    #put in blanks where there is not offgas data if shorter then inserted data lines.
                    if len(tank_file[tank_line]) < total_line_length:
                        tank_file[tank_line].extend(insertion_length_list)
            if header[2:] not in tank_file[0]:
                tank_file[0].extend(header[2:]) 
        return tank_file


    def combine_onlineOffgas_and_offline(self):
        print('Loading Offline Workbook')
        wb_offline = load_workbook(self.v3_file_text.get(), data_only=True)
        print('Workbook loaded')

        #get test offline test condition data.
        offline_tc = self.get_test_conditions(wb_offline)
        print('Grabbing test conditions.')

        #if offline eft time is > eft time in onilne data it wont be placed. #TODO add catch for this in future. 
        new_file_cb = []
        offline_data_headers  = []

        last_vessel = False
        offline_data = []

        for row in self.online_offgas_tank_file:
            #get the vessel tag. V1, V2, ect. 
            vessel_id = row[HC.online_id_column]
            try:
                v_location = vessel_id.index('V') #stop from breaking on header line.
            except ValueError:
                #put in offline  test condition headers and add data headers to new file being compiled.
                #TODO dont use error as condtion you fool. do better
                for position,tc_info in enumerate(["Strain", "Peptide","Test Condition"]):
                    row.insert(HC.online_tc_col_start+position,tc_info)
                new_file_cb.append(row)
                continue

            #pull out vessel tag from batch string. #may change if vessel column is added.
            vessel = vessel_id[v_location:v_location+2] #only works for tank numbers less than ten. #TODO fix for higher digits.
            if vessel != last_vessel:
                #check if there is any leftover offline data. Works only because tanks files are called in order. 
                if any(offline_data):
                    print(f'{offline_data} not placed. EFT time too short maybe?')

                #when the data changes to a new vessel run the code to pull the offline data.
                #this section could be pulled out.
                offline_vessel_sheet =wb_offline[vessel]
                print(f'Working on {vessel}')

                #Get where the data column ends on offline vessel sheet. Makes it dynamic to changing number of headers. 
                offline_column_end = len(offline_vessel_sheet[HC.offline_header_row_start])

                #Get a list of list with [eft, (data tuple)] here and rest of code assumes it is pulled in cronological order.
                offline_data = self.grab_offline_data(offline_vessel_sheet, HC.offline_eft_col,
                HC.offline_row_start, HC.offline_column_start, offline_column_end)    
                #save all the vessels headers to compare if they all match later. If so, add then to the top. 
                offline_vessel_header = []
                for headers in offline_vessel_sheet.iter_rows(HC.offline_header_row_start,
                HC.offline_header_row_end, HC.offline_header_column_start, offline_column_end, True):
                    for header in headers:
                        offline_vessel_header.append(header)
                offline_data_headers.append(offline_vessel_header)

            if any(offline_data):
                #check if the first eft goes in current row. if so place it and remove it from the list to check. doesn't check all, assumes cronological order.
                try:
                    if offline_data[0][0] <= float(row[HC.online_eft_index]):
                        for data_point in offline_data[0][1]:
                            row.append(data_point)
                        offline_data.pop(0) 
                except TypeError:
                    #TODO why is this here
                    print(f'A TypeError was thrown for {row[HC.online_eft_index]}. Placed in {row}')
                    for data_point in offline_data[0][1]:
                        row.append(data_point)
                    offline_data.pop(0) 
                    

            if any(row[5:]): #if row in empty after id, time, strain, ect. dont't add.
                #get list of test conditions from built dictionary. Add them in order into the row.
                test_condition =  offline_tc.get(vessel)
                if test_condition:
                    for position,tc_info in enumerate(test_condition):
                        row.insert(HC.online_tc_col_start+position,tc_info)
                    #add the new row to the new file being compiled. 
                    new_file_cb.append(row)

            #store the last vessel checked to see if new row is a different vessel.        
            last_vessel = vessel 

        wb_offline.close()
        print('Workbook closed')

        print('Checking if all Headers match.')
        if not all(head == offline_data_headers[0] for head in offline_data_headers):
            print("Offline file headers did not match up. Abort")
            return #TODO put stop the code and let user know the headers did not match up.

        #add in the headers to the first row. 
        for head in offline_data_headers[0]:
            new_file_cb[0].append(head)     

        return new_file_cb

    
    def get_test_conditions(self, offline_wb):
        #returns a a dictionary. With the key the vessel id and the value the [strain, peptide, testcondition]
        overview_worksheet = offline_wb["Setup"]  #TODO move hardcode to class object.
        overview_columns =  HC.offline_tc_col                            
        overview_rows =  HC.offline_tc_row
        test_info = {} 
        for x in overview_rows:
            row_info =[]
            for y in overview_columns[1:]:
                row_info.append(overview_worksheet.cell(x,y).value)
            #the id in SA files attaches run information before it in the cell. Pull out just the Vessel number. i.e. V2
            fermentor = overview_worksheet.cell(x,overview_columns[0]).value.strip()
            vessel= fermentor[fermentor.index("V"):fermentor.index("V")+2]   
            test_info[vessel] = row_info
        return test_info
    

    def grab_offline_data(self, work_sheet_offline, eft_column, eft_row, data_column_start, data_column_end):
        print('Pulling offline data')
        #this assumes that there is no break in the data column for eft.
        splice_counter = 0 #just to keep my dumb self from getting stuck in the infinite. 
        eft_row_end = eft_row
        offline_eft_list = []
        offline_eft_list.append([work_sheet_offline.cell(row=eft_row_end, column=eft_column).value]) 
        #build a list of all the offline eft times. and get the row that the data stops at.
        while work_sheet_offline.cell(row=eft_row_end, column=eft_column-1).value != None and splice_counter <= 40:   #eft_column-1 makes stoping row be based off data/time. Fixed code break when EFT has cells with infomation in them past the last row with tracked data.
            splice_counter +=1
            eft_row_end +=1 #this will step past the row by one. enters while loop on last row, adds one and then the while loop checks. None value so it breaks but the end row has been set to the empty row.
            if work_sheet_offline.cell(row=eft_row_end, column=eft_column-1).value != None: #TODO column=eft_column-1 needs to be its own so if headers change it wont break placement.  same as above above
                offline_eft_list.append([work_sheet_offline.cell(row=eft_row_end, column=eft_column).value])
            else:
                eft_row_end -=1
                break
        #get the data from the offline vessel sheet. Combine it with the list of eft times from above. 
        offline_data_tul = work_sheet_offline.iter_rows(eft_row, eft_row_end, data_column_start, data_column_end, True)
        offline_data_list =[]
        for tul in offline_data_tul:
            offline_data_list.append(tul)
        for eft_time in offline_eft_list:
            eft_time.append(offline_data_list[offline_eft_list.index(eft_time)])       
        
        #returns a list of list with [[eft, tuple of data for that eft], [next eft and tuple]...]
        #the tuple contains the complete data in whatever order it is in the offline sheet.
        #The order of the offline data and the order of the inserted headers are the same so there is no check that data is put in right column.
        #because the data is continous. I grabed it in one chunk that is hardcoded in the function input. 
        return offline_eft_list

    def get_file_path_xlsm(self, file_text):
        file_path = filedialog.askopenfilename(filetypes= [('Xlsm file', '.xlsm'),('CSV files','.csv'), ('Any','.*')])
        #tkinter set text on GUI
        file_text.set(file_path)
        return 


class Daily_Sample_Tab(Vessel_Select):
    def __init__(self, parent, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.daily_sample_bttn = ttk.Button(self, text='Daliy Sample Info', command=self.daily_sample)
        self.daily_sample_bttn.grid(row=1, column=0, padx=20, pady=5, sticky=tk.W)

        self.vessel_select_instructions = '''Pull computer recored data that is on the vessel run sheet\n\nEnter any number of run sheets exported from the Sartorius software.\n\nClick Daily Sample button\n\nYou will be asked for the run sheet'''
        self.vessel_info_text_id = self.vessel_info_canvas.create_text(20, 40, text=self.vessel_select_instructions, anchor=tk.NW, justify='center', width=180, font=('Times', '11'))

        self.vessel_select_frame.config( text = 'Vessel Run Data')

    def daily_sample(self):
        '''Get the sample data for the daily samples from the online data files.'''
        self.v1.columns_to_grab = [0, 2, 23, 45, 33, 29, 37, 9] #columns that have desired data and in correct order from csv file. 
        self.v2.columns_to_grab = [0, 2, 23, 45, 34, 30, 39, 9]
        self.v3.columns_to_grab = [0, 2, 19, 35, 25, 21, 27, 5]
        self.v4.columns_to_grab = [0, 2, 19, 35, 25, 23, 29, 5]
        self.v5.columns_to_grab = [0, 2, 21, 38, 29, 25, 32, 5]

        self.v1.batch_run_file = self.v1_file_text.get()
        self.v2.batch_run_file = self.v2_file_text.get()
        self.v3.batch_run_file = self.v3_file_text.get()
        self.v4.batch_run_file = self.v4_file_text.get()
        self.v5.batch_run_file = self.v5_file_text.get()

        tank_list = [tank for tank in [self.v1, self.v2, self.v3, self.v4, self.v5] if tank.batch_run_file.lower().endswith('.csv') ] 

        print('Select Run Sheet.')
        offline_run_wb =  load_workbook(filedialog.askopenfilename(filetypes= [('Xlsm files','.xlsm'),('CSV files','.csv')], title='Select Run Sheet'), data_only=True)

        #loop through tank list and call get eft
        for tank in tank_list:
            tank.eft_times = tank.grab_offline_eft_times(offline_run_wb[tank.runsheet_label])

        #set save file name and run through each data file and pull out eft time data. Assumes eft lists are in numerical order.
        save_file = filedialog.asksaveasfilename(filetypes= [('CSV files','.csv')], defaultextension='.csv')  
        with open(f'{save_file}', 'w', newline='') as file:
            print(f'Writing to file.')
            writer = csv.writer(file)
            for tank in tank_list:
                new_file = tank.get_daily_data()
                for row in new_file:
                    writer.writerow(row)
            print('File Saved')


def main():
    window = tk.Tk()
    window.title('SA File Combine')
    window['padx'] = 5
    window['pady'] = 40
    window.geometry('+450+250')
    window.columnconfigure(0, weight=1)
    MainApp(window)
    window.mainloop()

if __name__ == '__main__':
    main()
